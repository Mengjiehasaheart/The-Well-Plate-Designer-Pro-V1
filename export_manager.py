"""
Copyright (c) 2024 Mengjie Fan. All rights reserved.
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import json
import csv
from datetime import datetime
from typing import Dict, List, Optional
import io

class ExportManager:
    def __init__(self):
        self.border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
    def export_to_excel(self, plate: Dict, filename: Optional[str] = None) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = f"{plate['type']} Layout"
        
        ws['A1'] = f"Well Plate Designer Pro - {plate['type']}"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Created: {plate['created'].strftime('%Y-%m-%d %H:%M')}"
        ws['A2'].font = Font(size=10, italic=True)
        
        start_row = 4
        start_col = 2
        
        for j in range(plate['cols']):
            cell = ws.cell(row=start_row, column=start_col + j, value=j + 1)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)
        
        for i in range(plate['rows']):
            cell = ws.cell(row=start_row + i + 1, column=start_col - 1, value=chr(65 + i))
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)
        
        for i in range(plate['rows']):
            for j in range(plate['cols']):
                well_id = f"{chr(65 + i)}{j + 1}"
                well_info = plate['wells'][well_id]
                
                cell = ws.cell(row=start_row + i + 1, column=start_col + j)
                
                if well_info.get('treatment'):
                    # Check if this is a compound mixture
                    if well_info.get('compound_mixture'):
                        mixture_components = []
                        for comp in well_info['compound_mixture']:
                            mixture_components.append(f"{comp['compound']} ({comp['concentration']} {comp['unit']})")
                        cell.value = "\n".join(mixture_components)
                        if well_info.get('replicate'):
                            cell.value += f"\nRep {well_info['replicate']}"
                    else:
                        cell.value = f"{well_info.get('compound', '')}\n{well_info.get('treatment', '')}"
                        if well_info.get('replicate'):
                            cell.value += f"\nRep {well_info['replicate']}"
                    
                    color = well_info.get('color', '#FFFFFF').lstrip('#')
                    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    cell.fill = fill
                else:
                    cell.value = ""
                
                cell.border = self.border_style
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                ws.row_dimensions[start_row + i + 1].height = 40
                ws.column_dimensions[get_column_letter(start_col + j)].width = 15
        
        legend_row = start_row + plate['rows'] + 3
        ws.cell(row=legend_row, column=1, value="Legend:").font = Font(bold=True)
        
        unique_treatments = {}
        for well_info in plate['wells'].values():
            if well_info.get('treatment'):
                treatment = well_info['treatment']
                if treatment not in unique_treatments:
                    unique_treatments[treatment] = well_info.get('color', '#FFFFFF')
        
        legend_row += 1
        for treatment, color in unique_treatments.items():
            cell = ws.cell(row=legend_row, column=1, value=treatment)
            color_hex = color.lstrip('#')
            cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
            cell.border = self.border_style
            legend_row += 1
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def export_to_csv(self, plate: Dict) -> str:
        output = io.StringIO()
        writer = csv.writer(output)
        
        writer.writerow([f"Well Plate Designer Pro - {plate['type']}"])
        writer.writerow([f"Created: {plate['created'].strftime('%Y-%m-%d %H:%M')}"])
        writer.writerow([])
        
        header = [''] + [str(i+1) for i in range(plate['cols'])]
        writer.writerow(header)
        
        for i in range(plate['rows']):
            row = [chr(65 + i)]
            for j in range(plate['cols']):
                well_id = f"{chr(65 + i)}{j + 1}"
                well_info = plate['wells'][well_id]
                
                if well_info.get('treatment'):
                    content = f"{well_info.get('treatment', '')}"
                    if well_info.get('compound'):
                        content += f" - {well_info['compound']}"
                    if well_info.get('replicate'):
                        content += f" (Rep {well_info['replicate']})"
                    row.append(content)
                else:
                    row.append("")
            
            writer.writerow(row)
        
        writer.writerow([])
        writer.writerow(['Well Details:'])
        writer.writerow(['Well', 'Treatment', 'Compound', 'Subject', 'Replicate'])
        
        for well_id in sorted(plate['wells'].keys()):
            well_info = plate['wells'][well_id]
            if well_info.get('treatment'):
                # Handle compound mixtures
                if well_info.get('compound_mixture'):
                    mixture_str = " + ".join([f"{c['compound']} ({c['concentration']} {c['unit']})" 
                                            for c in well_info['compound_mixture']])
                    compound_value = mixture_str
                else:
                    compound_value = well_info.get('compound', '')
                
                writer.writerow([
                    well_id,
                    well_info.get('treatment', ''),
                    compound_value,
                    well_info.get('subject', ''),
                    well_info.get('replicate', '')
                ])
        
        return output.getvalue()
    
    def export_to_json(self, plate: Dict) -> str:
        export_data = {
            'metadata': {
                'type': plate['type'],
                'rows': plate['rows'],
                'cols': plate['cols'],
                'created': plate['created'].isoformat(),
                'version': '0.1.0',
                'generator': 'Well Plate Designer Pro'
            },
            'wells': {}
        }
        
        for well_id, well_info in plate['wells'].items():
            if well_info.get('treatment'):
                well_export = {
                    'treatment': well_info.get('treatment'),
                    'compound': well_info.get('compound'),
                    'subject': well_info.get('subject'),
                    'replicate': well_info.get('replicate'),
                    'color': well_info.get('color')
                }
                # Include compound mixture if present
                if well_info.get('compound_mixture'):
                    well_export['compound_mixture'] = well_info['compound_mixture']
                export_data['wells'][well_id] = well_export
        
        return json.dumps(export_data, indent=2)
    
    def import_from_json(self, json_data: str) -> Dict:
        try:
            data = json.loads(json_data)
            
            plate = {
                'type': data['metadata']['type'],
                'rows': data['metadata']['rows'],
                'cols': data['metadata']['cols'],
                'created': datetime.fromisoformat(data['metadata']['created']),
                'wells': {}
            }
            
            for i in range(plate['rows']):
                for j in range(plate['cols']):
                    well_id = f"{chr(65 + i)}{j + 1}"
                    if well_id in data['wells']:
                        well_data = data['wells'][well_id].copy()
                        # Ensure all fields are present
                        if 'compound_mixture' not in well_data:
                            well_data['compound_mixture'] = None
                        plate['wells'][well_id] = well_data
                    else:
                        plate['wells'][well_id] = {
                            'treatment': None,
                            'compound': None,
                            'compound_mixture': None,
                            'subject': None,
                            'replicate': None,
                            'color': '#2D3748'
                        }
            
            return plate
        except Exception as e:
            raise ValueError(f"Invalid JSON format: {str(e)}")
    
    def generate_report(self, plate: Dict) -> str:
        summary = []
        summary.append(f"# Well Plate Report")
        summary.append(f"**Type:** {plate['type']}")
        summary.append(f"**Created:** {plate['created'].strftime('%Y-%m-%d %H:%M')}")
        summary.append(f"**Dimensions:** {plate['rows']} × {plate['cols']}")
        summary.append("")
        
        total_wells = plate['rows'] * plate['cols']
        assigned_wells = sum(1 for w in plate['wells'].values() if w.get('treatment'))
        
        summary.append(f"## Statistics")
        summary.append(f"- Total wells: {total_wells}")
        summary.append(f"- Assigned wells: {assigned_wells}")
        summary.append(f"- Empty wells: {total_wells - assigned_wells}")
        summary.append(f"- Utilization: {assigned_wells/total_wells*100:.1f}%")
        summary.append("")
        
        treatments = {}
        for well_info in plate['wells'].values():
            if well_info.get('treatment'):
                treatment = well_info['treatment']
                if treatment not in treatments:
                    treatments[treatment] = 0
                treatments[treatment] += 1
        
        if treatments:
            summary.append(f"## Treatment Distribution")
            for treatment, count in sorted(treatments.items()):
                summary.append(f"- {treatment}: {count} wells")
        
        return "\n".join(summary)
    
    def export_to_excel_long_format(self, plate: Dict) -> bytes:
        """Export plate data in long format for data science workflows"""
        # Create a list to store all well data
        data_rows = []
        
        # Extract plate metadata
        plate_type = plate['type']
        plate_rows = plate['rows']
        plate_cols = plate['cols']
        created_date = plate['created'].strftime('%Y-%m-%d %H:%M')
        
        # Iterate through all wells in order
        for i in range(plate['rows']):
            for j in range(plate['cols']):
                well_id = f"{chr(65 + i)}{j + 1}"
                well_info = plate['wells'][well_id]
                
                # Create base row data
                row_data = {
                    'plate_type': plate_type,
                    'plate_rows': plate_rows,
                    'plate_cols': plate_cols,
                    'created_date': created_date,
                    'well_id': well_id,
                    'row': chr(65 + i),
                    'column': j + 1,
                    'row_index': i,
                    'column_index': j,
                }
                
                # Add treatment information
                if well_info.get('treatment'):
                    row_data['treatment'] = well_info.get('treatment', '')
                    row_data['compound'] = well_info.get('compound', '')
                    row_data['subject'] = well_info.get('subject', '')
                    row_data['replicate'] = well_info.get('replicate', '')
                    row_data['color'] = well_info.get('color', '')
                    
                    # Handle compound mixtures
                    if well_info.get('compound_mixture'):
                        # Create a row for each component in the mixture
                        for idx, component in enumerate(well_info['compound_mixture']):
                            mixture_row = row_data.copy()
                            mixture_row['mixture_component'] = idx + 1
                            mixture_row['mixture_compound'] = component['compound']
                            mixture_row['mixture_concentration'] = component['concentration']
                            mixture_row['mixture_unit'] = component['unit']
                            data_rows.append(mixture_row)
                    else:
                        # Single compound - add additional fields
                        row_data['mixture_component'] = None
                        row_data['mixture_compound'] = None
                        row_data['mixture_concentration'] = None
                        row_data['mixture_unit'] = None
                        
                        # Extract concentration if present
                        if well_info.get('concentration'):
                            row_data['concentration'] = well_info.get('concentration', '')
                        
                        # Extract time point if present
                        if well_info.get('time_point'):
                            row_data['time_point'] = well_info.get('time_point', '')
                        
                        data_rows.append(row_data)
                else:
                    # Empty well
                    row_data['treatment'] = None
                    row_data['compound'] = None
                    row_data['subject'] = None
                    row_data['replicate'] = None
                    row_data['color'] = None
                    row_data['mixture_component'] = None
                    row_data['mixture_compound'] = None
                    row_data['mixture_concentration'] = None
                    row_data['mixture_unit'] = None
                    row_data['concentration'] = None
                    row_data['time_point'] = None
                    data_rows.append(row_data)
        
        # Convert to DataFrame
        df = pd.DataFrame(data_rows)
        
        # Reorder columns for better readability
        column_order = [
            'plate_type', 'well_id', 'row', 'column', 'row_index', 'column_index',
            'treatment', 'compound', 'concentration', 'time_point', 'subject', 'replicate',
            'mixture_component', 'mixture_compound', 'mixture_concentration', 'mixture_unit',
            'color', 'plate_rows', 'plate_cols', 'created_date'
        ]
        
        # Only include columns that exist
        columns_to_include = [col for col in column_order if col in df.columns]
        df = df[columns_to_include]
        
        # Create Excel file
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Plate_Data_Long_Format', index=False)
            
            # Access the worksheet to apply formatting
            worksheet = writer.sheets['Plate_Data_Long_Format']
            
            # Format headers
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                
                # Auto-adjust column width
                column_letter = get_column_letter(col_idx)
                max_length = max(
                    len(str(col_name)),
                    df[col_name].astype(str).str.len().max() if not df[col_name].isna().all() else 0
                )
                worksheet.column_dimensions[column_letter].width = min(max_length + 2, 30)
            
            # Add borders to all cells
            for row in worksheet.iter_rows(min_row=1, max_row=len(df) + 1, 
                                         min_col=1, max_col=len(df.columns)):
                for cell in row:
                    cell.border = self.border_style
            
            # Add color formatting to cells based on treatment color
            if 'color' in df.columns:
                for idx, row in df.iterrows():
                    if row['color'] and not pd.isna(row['color']):
                        color_hex = row['color'].lstrip('#')
                        fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
                        # Apply color to the well_id column
                        well_id_col = df.columns.get_loc('well_id') + 1
                        worksheet.cell(row=idx + 2, column=well_id_col).fill = fill
        
        output.seek(0)
        return output.getvalue()